from __future__ import print_function, unicode_literals

import os
import re
import math
import glob

import click
import polib
from openpyxl import load_workbook


def hr(char='\u2500', width=None, **kwargs):
    if width is None:
        width = click.get_terminal_size()[0]
    mult = int(math.ceil(width * 1.0 / len(char)))
    click.secho((char * mult)[:width], **kwargs)


def get_pofiles(locale_folders, language):
    for folder in locale_folders:
        rw_ro, folder = folder.split(':', 1)
        assert rw_ro in ['rw', 'ro']
        read_only = rw_ro == 'ro'
        po_file = os.path.join(folder, language, 'LC_MESSAGES', '*.po')
        for path in glob.glob(po_file):
            yield read_only, polib.pofile(path, wrapwidth=78)


class MsgStr(object):
    def __init__(self, entry, multiplicity=0):
        self.entry = entry
        self.multiplicity = multiplicity

    def get(self):
        if self.entry.msgid_plural:
            return self.entry.msgstr_plural[self.multiplicity]
        else:
            assert self.multiplicity == 0
            return self.entry.msgstr

    def normalize_whitespace(self, value):
        if value is None:
            return
        msgid = self.entry.msgid
        if self.entry.msgid_plural and self.multiplicity:
            msgid = self.entry.msgid_plural
        space_before = re.search(r'^\s*', msgid).group(0)
        space_after = re.search(r'\s*$', msgid).group(0)
        return ''.join([space_before, value.strip(), space_after])

    def set(self, value):
        if self.entry.msgid_plural:
            self.entry.msgstr_plural[self.multiplicity] = value
        else:
            self.entry.msgstr = value


@click.command()
@click.option('--skip', type=int, default=0)
@click.option('--ctx-col', '-c', type=int, default=2)
@click.option('--key-col', '-k', type=int, default=3)
@click.option('--trans-col', '-t', type=int, default=4)
@click.option('--sheet', '-s', type=int, default=1)
@click.option('--pretend/--no-pretend', '-p', default=False)
@click.option('--language', '-l')
@click.option('--add/--no-add', '-a')
@click.option('--hide-ok/--show-ok', default=False)
@click.argument('workbook', type=click.Path(exists=True))
@click.argument('locale_folders', nargs=-1)
def main(workbook, locale_folders, language, key_col, trans_col, sheet,
         pretend, hide_ok, ctx_col, skip, add):
    wb = load_workbook(workbook)
    ws = list(wb)[sheet-1]
    rows = iter(ws.iter_rows())

    for i in range(skip):
        next(rows)
    headers = next(rows)

    if key_col is None or trans_col is None:
        click.echo('The following columns are available:')
        for i, cell in enumerate(headers):
            if cell.value is not None:
                click.echo('{:2d}. {}'.format(i + 1, cell.value))

        if key_col is None:
            key_col = click.prompt('Which column contains the key?',
                                   type=int, default=1)

        if trans_col is None:
            trans_col = click.prompt('Which column contains the translation?',
                                     type=int)

    ctx_col -= 1
    key_col -= 1
    trans_col -= 1

    hr(fg='yellow')
    click.secho(' Selected sheet: {}'.format(ws.title))
    click.secho(' Key column:     {}'.format(headers[key_col].value))
    click.secho(' Value column:   {}'.format(headers[trans_col].value))
    click.secho(' Language:       {}'.format(language))
    hr(fg='yellow')

    pofiles = list(get_pofiles(locale_folders, language))

    for row in rows:
        ctx = row[ctx_col].value or None
        key = row[key_col].value
        plr = row[key_col + 1].value
        trans = row[trans_col].value

        # Check flag
        flag = int(row[trans_col - 1].value)
        fuzzy = (flag == 1)

        if key is None and trans is None:
            # This row can be safely ignored
            continue

        entry, msgstr, read_only = find_entry(pofiles, ctx, key, plr, trans)
        if entry:
            handle_entry(entry, msgstr, ctx, key, trans, fuzzy, hide_ok)
        elif plr == 's':
            # TODO: Support adding pluralized entries as well
            if add or click.confirm('Do you want to add it to the file?'):
                entry = polib.POEntry(
                    msgid=key or '',
                    msgstr=trans or '',
                    msgctx=ctx or None,
                )
                for read_only, po in pofiles:
                    if not read_only:
                        po.append(entry)
                        break

    if not pretend:
        for read_only, po in pofiles:
            if not read_only:
                po.save()


def find(pofile, st, by='msgid', include_obsolete_entries=False,
         msgctxt=False):
    if include_obsolete_entries:
        entries = pofile[:]
    else:
        entries = [e for e in pofile if not e.obsolete]
    for e in entries:
        if getattr(e, by).strip() == st:
            if msgctxt is not False and e.msgctxt != msgctxt:
                continue
            return e
    return None


def find_entry(pofiles, ctx, key, plr, trans):
    msgstr = None
    args_set = [
        {'include_obsolete_entries': False},
        {'include_obsolete_entries': True},
    ]

    for kwargs in args_set:
        if plr == 'P':
            kwargs['by'] = 'msgid_plural'
        for read_only, po in pofiles:
            entry = find(po, key, msgctxt=ctx, **kwargs)
            if entry is not None:
                msgstr = MsgStr(entry, plr == 'P')
                break
        if msgstr:
            break
    else:
        # TODO: How do we want to handle this?
        key = '{}:{}'.format(ctx, key) if ctx else key
        click.secho('The entry "{}" was not found.'.format(key), fg='red')
        return None, None, False

    if read_only:
        if trans != msgstr.get():
            key = '{}:{}'.format(ctx, key)
            click.secho('The entry "{}" has been translated as "{}" in "{}" '
                        'but the file is marked as read-only.'
                        .format(key, msgstr.get(), po.fpath), fg='red')
        return None, None, True

    return entry, msgstr, False


def handle_entry(entry, msgstr, ctx, key, trans, fuzzy, hide_ok):
    trans_norm = msgstr.normalize_whitespace(trans)

    if not trans or trans.strip() == msgstr.get().strip():
        # Already up to date
        if not hide_ok:
            click.secho('The entry "{}" is already up to date.'.format(key),
                        fg='green')
        if msgstr.get().strip():
            msgstr.set(msgstr.normalize_whitespace(msgstr.get()))
        else:
            msgstr.set('')
        if 'fuzzy' in entry.flags and not fuzzy:
            entry.flags.remove('fuzzy')
        entry.obsolete = False
        return

    if not entry.translated():
        # Entry is in the file, but has not been translated
        click.secho('The entry "{}" has not been translated yet.'
                    .format(key), fg='yellow')
        msgstr.set(trans_norm)
        if 'fuzzy' in entry.flags and not fuzzy:
            entry.flags.remove('fuzzy')
        # entry.obsolete = False
        return

    # Entry has been translated but needs update
    # TODO: Ask for confirmation
    hr()
    click.secho(str(ctx))
    click.secho('The entry "{}" was updated:'.format(key))
    click.secho(' OLD: "{}"'.format(msgstr.get()))
    click.secho(' NEW: "{}"'.format(trans_norm))
    msgstr.set(trans)
    hr()
