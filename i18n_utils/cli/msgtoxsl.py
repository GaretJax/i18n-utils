from __future__ import print_function, unicode_literals

import collections
from textwrap import wrap

import click

from openpyxl import Workbook
from openpyxl.cell import get_column_letter

from i18n_utils import styles, params


@click.command()
@click.option('--locale', '-l', 'locales', multiple=True)
@click.option('--key-locale', '-k', default='en')
@click.option('--title', '-t', default='My project')
@click.argument('out', type=click.Path(exists=False, dir_okay=False,
                                       writable=True))
@click.argument('folders', nargs=-1, type=params.LocaleFolderParamType())
def main(locales, key_locale, title, folders, out):
    wb = Workbook()

    if not locales:
        locales = list(folders[0].locales())

    add_translations_wb(wb.active, folders, locales, key_locale, title)
    wb.save(out)


def coord(col, row):
    if str(col).isdigit():
        col = get_column_letter(int(col))
    return '{}{}'.format(col, row)


def num_lines(s, chars=50):
    lines = s.split('\n')
    return sum(len(wrap(l, chars)) for l in lines)


def add_row(ws, left, context, key, key_idx, entries, plural=False):
    if plural:
        def get_str(entry):
            return entry.msgstr_plural[1].strip()
    else:
        def get_str(entry):
            if entry.msgid_plural:
                return entry.msgstr_plural[0].strip()
            else:
                return entry.msgstr.strip()

    pluralized = plural
    if not pluralized:
        for e in entries:
            if e:
                pluralized = bool(e.msgid_plural)
                break

    ws.append([''] * (len(entries) * 2 + 5 + left))
    row_idx = len(ws.rows)
    row = ws.rows[-1][left:]

    occurrences = sum((entry.occurrences for entry in entries if entry), [])
    occurrences_dict = collections.defaultdict(lambda: set())
    for path, line in occurrences:
        occurrences_dict[path].add(line)

    def fmt(o, lines):
        return '{}:{}'.format(o, ','.join(sorted(lines)))

    occurrences = '\n'.join(fmt(*o) for o in sorted(occurrences_dict.items()))

    # Set row height
    content = [context or '', key] + [get_str(e) for e in entries if e]
    lines = max(num_lines(v) for v in content)
    height = (lines + 1) * styles.line_height
    height = max(height, 11 * num_lines(occurrences, 200) + 3)
    ws.row_dimensions[row_idx].height = height

    # Add context cell
    ctx_cell = row[0]
    ctx_cell.value = context
    ctx_cell.style = styles.row_header(is_odd=row_idx % 2, is_last=False)

    # Add key cell
    key_cell = row[1]
    key_cell.value = key
    key_cell.style = styles.row_header(is_odd=row_idx % 2, is_last=False)

    # Add plural cell
    plr_cell = row[2]
    plr_cell.value = 'P' if plural else 'p' if pluralized else 's'
    plr_cell.style = styles.row_header(is_odd=row_idx % 2, is_last=True)

    # Add translations cells
    row = row[3:]
    for i, entry in enumerate(entries):
        cb_cell = row[i * 2]
        msgstr = get_str(entry) if entry else ''
        if (entry and msgstr and 'fuzzy' not in entry.flags
                and not entry.obsolete):
            msg = msgstr
            cb_cell.value = 2
        elif entry and msgstr and 'fuzzy' in entry.flags:
            msg = msgstr
            cb_cell.value = 1
        else:
            if i == key_idx:
                msg = key
                cb_cell.value = 2
            else:
                msg = ''
                cb_cell.value = 0

        cb_cell.style = styles.translation_cb_cell(is_odd=row_idx % 2)

        trans_cell = row[i * 2 + 1]
        trans_cell.value = msg
        trans_cell.style = styles.translation_cell(is_odd=row_idx % 2)

    # Add comment
    cmt_cell = row[-2]
    cmt_cell.style = styles.comment_cell(is_odd=row_idx % 2)

    # Occurrences
    occ_cell = row[-1]
    occ_cell.style = styles.occurrences_cell(is_odd=row_idx % 2)
    occ_cell.value = occurrences


def add_translations_wb(ws, folders, locales, key_locale, title):
    top = 4
    left = 1

    ws.title = 'Translations'

    # Parse the PO files
    entries = collections.defaultdict(lambda: [None] * len(locales))

    for folder in folders:
        po_files = [(locale, po) for locale, po in folder.pofiles()
                    if locale in locales]

        for locale, po in po_files:
            click.secho('Parsing {}'.format(po.fpath), fg='yellow')
            idx = locales.index(locale)
            for entry in po:
                if entry.obsolete:
                    continue
                if locale == key_locale and not entry.msgstr:
                    continue

                key = entry.msgctxt, entry.msgid.strip(), False

                try:
                    current_entry = entries[key][idx]
                except (KeyError, IndexError):
                    pass
                else:
                    if current_entry:
                        if current_entry.msgid_plural:
                            current_entry.msgstr_plural[0].strip()
                        else:
                            msgstr = current_entry.msgstr.strip()

                        if msgstr:
                            continue

                entries[key][idx] = entry

                # Handle plurals
                if not entry.msgid_plural:
                    continue

                key = entry.msgctxt, entry.msgid_plural.strip(), True

                try:
                    current_entry = entries[key][idx]
                except (KeyError, IndexError):
                    pass
                else:
                    if current_entry:
                        if current_entry.msgstr_plural[1].strip():
                            continue

                entries[key][idx] = entry

    entries = sorted(entries.items(), key=lambda e: (e[0][0], e[0][1].lower()))
    key_idx = locales.index(key_locale)

    # Add the translations
    # ws = wb.create_sheet()
    ws.sheet_view.showGridLines = False

    # Sheet title ############################################################
    header = ws.cell(coord(left + 1, top - 2))
    header.value = 'Translations for "{}"'.format(title)
    header.style = styles.title_style

    # Headers ################################################################
    ctx_cell = ws.cell(coord(left + 1, top))
    ctx_cell.style = styles.column_header()
    ctx_cell.value = 'Context'

    ws.merge_cells('{}:{}'.format(coord(left + 2, top), coord(left + 3, top)))
    key_cell = ws.cell(coord(left + 2, top))
    key_cell.style = styles.column_header()
    key_cell.value = 'Key ({} entries)'.format(len(entries))
    ws.cell(coord(left + 3, top)).style = styles.column_header()

    for i, l in enumerate(locales):
        from_cell = coord(i * 2 + 4 + left, top)
        to_cell = coord(i * 2 + 4 + left + 1, top)
        ws.merge_cells('{}:{}'.format(from_cell, to_cell))
        locale_header_text = (
            '=CONCATENATE('
            '    "{name} (",'
            '    ROUND(COUNTIF({start}:{end}, "=2")/{count} *100, 0),'
            '    "% translated)"'
            ')'
        )
        cell = ws.cell(from_cell)
        cell.style = styles.column_header()
        cell.value = locale_header_text.format(
            name=l.upper(),
            start=coord(i * 2 + 4 + left, top + 1),
            end=coord(i * 2 + 4 + left, len(entries) + top + 1),
            count=len(entries)
        )
        ws.cell(to_cell).style = styles.column_header()

    cmt_cell = ws.cell(coord(len(locales) * 2 + 4 + left, top))
    cmt_cell.style = styles.column_header()
    cmt_cell.value = 'Comment'

    occ_cell = ws.cell(coord(len(locales) * 2 + 5 + left, top))
    occ_cell.style = styles.column_header()
    occ_cell.value = 'Occurrences'

    # Add translations #######################################################
    for i, ((ctx, key, plural), translations) in enumerate(entries):
        add_row(ws, left, ctx, key, key_idx, translations, plural=plural)

    # Column widths ##########################################################

    # Left offset
    for i in range(left):
        ws.column_dimensions[get_column_letter(i + 1)].width = 5

    # Context and key
    ws.column_dimensions[get_column_letter(left + 1)].width = 30
    ws.column_dimensions[get_column_letter(left + 2)].width = 50
    ws.column_dimensions[get_column_letter(left + 3)].width = 5

    # Translations
    for i in range(len(locales)):
        col = left + 4 + i * 2
        ws.column_dimensions[get_column_letter(col)].width = 5
        ws.column_dimensions[get_column_letter(col + 1)].width = 50

    # Comment
    col = left + 4 + len(locales) * 2
    ws.column_dimensions[get_column_letter(col)].width = 100

    # Occurrences
    col = left + 5 + len(locales) * 2
    ws.column_dimensions[get_column_letter(col)].width = 80

    # Row heights ############################################################
    for i in range(top):
        ws.row_dimensions[i + 1].height = 25
    ws.row_dimensions[top].height = 40

    # Conditional formatting #################################################
    for i in range(len(locales)):
        from_cell = coord(left + 4 + i * 2, top + 1)
        to_cell = coord(left + 4 + i * 2, top + 1 + len(entries))
        ws.conditional_formatting.add(
            '{}:{}'.format(from_cell, to_cell),
            {
                'iconSet': {'cfvo': [{'type': 'num', 'val': '0'},
                                     {'type': 'num', 'val': '1'},
                                     {'type': 'num', 'val': '2'}],
                            'iconSet': '3Flags',
                            'showValue': '0'},
                'priority': 1,
                'type': 'iconSet'
            }
        )

    # Freeze and protect #####################################################
    ws.freeze_panes = ws[coord(left + 4, top + 1)]
    ws.protection.enable()
