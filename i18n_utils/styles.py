from openpyxl.styles import Font, Style, Alignment, Border, Side, PatternFill
from openpyxl.styles import borders, fills
from openpyxl.styles import Protection

from .utils import memoize


cell_fills = [
    PatternFill(
        fill_type=fills.FILL_SOLID,
        start_color='E6EFD7',
    ),
    PatternFill(
        fill_type=fills.FILL_SOLID,
        start_color='FFFFFF',
    ),
]
font = Font(name='Tahoma')
alignment = Alignment(
    vertical='center',
    shrink_to_fit=True,
    wrap_text=True,
    indent=1,
)
border = Side(border_style=borders.BORDER_THIN, color='B8D08A')
header_border = line_height = 13
title_style = Style(
    font=Font(name='Tahoma', size=20, bold=True)
)


@memoize
def column_header():
    return Style(
        font=Font(
            name='Thaoma',
            bold=True,
            color='FFFFFF',
            size=14,
        ),
        border=Border(
            left=Side(border_style=borders.BORDER_THIN, color='8BB048'),
            right=Side(border_style=borders.BORDER_THIN, color='8BB048'),
            top=Side(border_style=borders.BORDER_THIN, color='8BB048'),
            bottom=Side(border_style=borders.BORDER_THIN, color='8BB048'),
        ),
        alignment=Alignment(
            vertical='center',
            horizontal='center',
        ),
        fill=PatternFill(
            fill_type=fills.FILL_SOLID,
            start_color='8BB048',
        ),
    )


@memoize
def row_header(is_odd, is_last):
    return Style(
        font=font,
        alignment=Alignment(
            vertical='center',
            horizontal='center' if is_last else 'left',
            shrink_to_fit=True,
            wrap_text=True,
            indent=int(not is_last),
        ),
        border=Border(
            left=border,
            right=(Side(border_style=borders.BORDER_DOUBLE, color='8BB048')
                   if is_last else border),
            top=border,
            bottom=border,
        ),
        fill=cell_fills[int(bool(is_odd))],
        protection=Protection(),
    )


@memoize
def translation_cell(is_odd):
    return Style(
        font=font,
        alignment=Alignment(
            vertical='center',
            shrink_to_fit=True,
            wrap_text=True,
        ),
        border=Border(
            right=border,
            top=border,
            bottom=border,
        ),
        fill=cell_fills[int(bool(is_odd))],
        protection=Protection(locked=False),
    )


@memoize
def translation_cb_cell(is_odd):
    return Style(
        font=font,
        alignment=Alignment(
            vertical='center',
            horizontal='center',
        ),
        border=Border(
            left=border,
            top=border,
            bottom=border,
        ),
        fill=cell_fills[int(bool(is_odd))],
        protection=Protection(locked=False),
    )


comment_cell = translation_cell


@memoize
def occurrences_cell(is_odd):
    return Style(
        font=Font(name='Tahoma', size=8),
        alignment=Alignment(
            vertical='top',
            wrap_text=True,
        ),
        border=Border(
            right=border,
            top=border,
            bottom=border,
        ),
        fill=cell_fills[int(bool(is_odd))],
        protection=Protection(locked=False),
    )
